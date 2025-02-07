from flask import Flask, render_template, request, jsonify
import qrcode
import os
import datetime
import openpyxl

app = Flask(__name__)

# ✅ Define the path for attendance file
EXCEL_FILE = "attendance.xlsx"

def create_excel_file():
    """Create attendance.xlsx if it does not exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Teacher Name", "Date", "Time", "Year", "Day", "Timestamp", "Status"])
        wb.save(EXCEL_FILE)
        print("✅ Created 'attendance.xlsx' with headers.")

    # ✅ Force remove "Read-Only" mode
    try:
        os.chmod(EXCEL_FILE, 0o777)  # Full read/write access
    except Exception as e:
        print(f"⚠️ Could not change file permissions: {e}")

# ✅ Run this function when the app starts
create_excel_file()

# ✅ Updated Teacher List
VALID_TEACHERS = [
    "Daniel", "Nalina", "Johniel", "Mamtha", "Meena", "Asha", "Mohan", "Nagaraj", "Sunitha",
    "Viraj", "Akhil", "Bindiya", "Manjula", "Afiga", "Priyanka", "Shruti", "Subhashini",
    "Vieenitha", "Supriya", "Padmini", "Pushpajali", "Ashiga", "Grace", "Kavitha", "Kavitha VG",
    "Ammurutha", "Anannya", "Amurutha N", "Sumritha", "Jhansi Rani", "Sujatha", "Sangeeta",
    "Archana"
]

# ✅ Generate QR codes
qr_folder = "static/qrcodes"
os.makedirs(qr_folder, exist_ok=True)

for name in VALID_TEACHERS:
    qr_path = os.path.join(qr_folder, f"{name}.png")
    if not os.path.exists(qr_path):
        qr = qrcode.make(name)
        qr.save(qr_path)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/scan', methods=['POST'])
def scan_qr():
    data = request.json
    teacher_name = data.get("name")

    if teacher_name in VALID_TEACHERS:
        now = datetime.datetime.now()
        date = now.strftime("%Y-%m-%d")    
        time = now.strftime("%H:%M:%S")    
        year = now.strftime("%Y")          
        day = now.strftime("%A")           
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")  
        status = "Present"

        # ✅ Open and Save Attendance in Excel
        create_excel_file()  # Ensure file exists before saving
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([teacher_name, date, time, year, day, timestamp, status])
            wb.save(EXCEL_FILE)
            return jsonify({"message": f"✅ Attendance Saved: {teacher_name}", "timestamp": timestamp})
        except Exception as e:
            return jsonify({"message": f"⚠️ Error Saving: {str(e)}"}), 500
    
    return jsonify({"message": "❌ Invalid QR Code"}), 400

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/check_password', methods=['POST'])
def check_password():
    data = request.json
    password = data.get("password")

    if password == "noes1236":
        create_excel_file()  # Ensure file exists
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        attendance_records = [
            {"name": row[0], "date": row[1], "time": row[2], "year": row[3], 
             "day": row[4], "timestamp": row[5], "status": row[6]}
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        return jsonify({"success": True, "records": attendance_records})
    
    return jsonify({"success": False, "message": "❌ Incorrect Password"}), 403

if __name__ == '__main__':
    app.run(debug=True)
